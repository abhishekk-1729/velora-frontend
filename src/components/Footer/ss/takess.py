import openpyxl
import requests

# Define the API key and base URL
API_KEY = "key_bA6888muPUHreU6ptLYZaF"
# apiKey=key_hzf5A3cv7MYidVkexpPG5K
# apiKey=key_cNjXLviPtshsvaYMCkH5ED
# key_apUQQV1fyjfdfd5Y6X9uRA
# apiKey=key_bA6888muPUHreU6ptLYZaF
API_BASE_URL = "https://api.screenshotapi.com/take"

def get_screenshot_url(url):
    """
    Fetch the screenshot URL for a given website URL.
    """
    params = {
        "url": url,
        "apiKey": API_KEY,
        "viewportHeight": 300
    }
    try:
        response = requests.get(API_BASE_URL, params=params)
        response.raise_for_status()
        data = response.json()
        return data.get("outputUrl")
    except Exception as e:
        print(f"Error fetching screenshot for {url}: {e}")
        return None

def process_excel_file(input_file, output_file):
    """
    Process the Excel file to add screenshot URLs in column C.
    """
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Iterate through rows, starting from the second row (assuming headers in row 1)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        website_cell = row[0]  # Column B contains the website URLs
        screenshot_url_cell = sheet.cell(row=website_cell.row, column=3)  # Column C for screenshot URLs

        if website_cell.value:
            print(f"Processing: {website_cell.value}")
            screenshot_url = get_screenshot_url(website_cell.value)
            if screenshot_url:
                screenshot_url_cell.value = screenshot_url

    # Save the updated workbook
    wb.save(output_file)
    print(f"Updated Excel file saved as: {output_file}")

# Example usage
input_file = "rem.xlsx"  # Replace with your input file name
output_file = "websites_with_screenshots7.xlsx"  # Output file name
process_excel_file(input_file, output_file)