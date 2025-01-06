import os
import pandas as pd
import json
from openpyxl import load_workbook

def update_excel_with_json(input_excel, json_folder, output_excel):
    """
    Updates an Excel file with data from JSON files and saves the updated version.

    Parameters:
    - input_excel (str): Path to the input Excel file.
    - json_folder (str): Path to the folder containing JSON files.
    - output_excel (str): Path to save the updated Excel file.
    """
    # Load the Excel file
    workbook = load_workbook(input_excel)
    sheet = workbook.active

    # Start updating from the second row (assuming the first row is headers)
    for row_index in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_index, column=27).value
        json_filename = f"{int(cell_value)}.json" if isinstance(cell_value, (int, float)) else f"{cell_value}.json"
        json_path = os.path.join(json_folder, json_filename)

        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as json_file:
                response_json = json.load(json_file)

            # Update the sheet with data from the JSON file
            try:
                sheet.cell(row=row_index, column=3).value = response_json.get("hotelBookingAvailable", "N/A")

                contact_numbers = response_json.get("contactNumbers", ["N/A"])
                sheet.cell(row=row_index, column=4).value = ", ".join(contact_numbers)

                emails = response_json.get("emails", ["N/A"])
                sheet.cell(row=row_index, column=5).value = ", ".join(emails)

                sheet.cell(row=row_index, column=6).value = response_json.get("mobileResponsivenessOutOf10", "N/A")
                sheet.cell(row=row_index, column=7).value = response_json.get("mobileResponsivenessTip", "")

                sheet.cell(row=row_index, column=8).value = response_json.get("loadingSpeedOutOf10", "N/A")
                sheet.cell(row=row_index, column=9).value = response_json.get("loadingSpeedTip", "")

                sheet.cell(row=row_index, column=10).value = response_json.get("seoOptimizationOutOf10", "N/A")
                sheet.cell(row=row_index, column=11).value = response_json.get("seoOptimizationTip", "")

                sheet.cell(row=row_index, column=12).value = response_json.get("userExperienceOutOf10", "N/A")
                sheet.cell(row=row_index, column=13).value = response_json.get("userExperienceTip", "")

                sheet.cell(row=row_index, column=14).value = response_json.get("bookingFlowOutOf10", "N/A")
                sheet.cell(row=row_index, column=15).value = response_json.get("bookingFlowTip", "NA")

                tips = response_json.get("tips", ["No tips available"])
                sheet.cell(row=row_index, column=16).value = ", ".join(tips)

                sheet.cell(row=row_index, column=17).value = response_json.get("whatThisWebsiteDoesInOneLine", "N/A")

                sheet.cell(row=row_index, column=18).value = response_json.get("isSiteFunctional", "N/A")
                sheet.cell(row=row_index, column=19).value = response_json.get("belongsToThisHotel", "N/A")
                sheet.cell(row=row_index, column=20).value = response_json.get("sameDomainBooking", "N/A")
                sheet.cell(row=row_index, column=21).value = response_json.get("thirdPartyBookingDomain", "N/A")
                sheet.cell(row=row_index, column=22).value = response_json.get("category", "N/A")
                sheet.cell(row=row_index, column=23).value = response_json.get("ifOthersThenWhat", "N/A")

            except Exception as e:
                print(f"Error updating row {row_index} with file {json_filename}: {e}")
        else:
            print(f"JSON file not found for row {row_index}: {json_filename}")

    # Save the updated workbook
    workbook.save(output_excel)
    print(f"Updated Excel file saved as {output_excel}")

# Example usage
input_excel = "/Users/abhishekkumar/Downloads/final3.xlsx"  # Replace with your input Excel file path
json_folder = "/Users/abhishekkumar/velora-frontend/vite-project/src/components/Footer/final_scripts/results3"  # Replace with your JSON folder path
output_excel = "output_final3.xlsx"  # Replace with your desired output file path

update_excel_with_json(input_excel, json_folder, output_excel)
